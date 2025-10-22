








import pandas as pd
import json
import numpy as np
import openpyxl as pyxl
import re
import os
import shutil
import datetime



def check_unit_type(unit_type):
  # unit_type is a string denoting
  if isinstance(unit_type, str):
    unit_type = unit_type.strip()
    if unit_type in ["cs", "CS", "case"]:
      unit_type = "CASE"
    elif unit_type in ["lb", "lbs", "LBS"]:
      unit_type = "LB"
    elif unit_type in ["ea", "each","Each"]:
      unit_type = "EACH"
    elif unit_type in ["half"]:
      unit_type = "HALF"

  return unit_type





def move_and_archive_document(filename, origin_dir, destination_dir):
    """
    Copies a specified file from the origin directory to the destination directory,
    appending a datetime stamp to the copied file's name. The original file
    remains in the origin directory.

    Args:
        filename (str): The name of the file to move (e.g., 'invoice_1234.pdf').
        origin_dir (str): The full path to the directory where the file currently resides.
                          Example: 'vendors/sysco/inputs'
        destination_dir (str): The full path to the directory where the file should be moved.
                               Example: 'inputs/processed'
    """
    
    # --- 1. PREPARE PATHS AND TIMESTAMP ---

    # Get the current datetime in a safe format (e.g., YYYYMMDD_HHMMSS)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Separate the filename into name and extension
    name, ext = os.path.splitext(filename)
    
    # Create the new filename: name_timestamp.ext
    new_filename = f"{name}_{timestamp}{ext}"

    # Construct the full source path (the original file)
    source_path = os.path.join(origin_dir, filename)

    # Construct the full destination path with the new filename
    destination_path = os.path.join(destination_dir, new_filename)


    # --- 2. ENSURE DESTINATION DIRECTORY EXISTS ---
    if not os.path.exists(destination_dir):
        # Create destination and any necessary parent directories
        os.makedirs(destination_dir, exist_ok=True)
        print(f"Created destination directory: {destination_dir}")

    # --- 3. COPY THE FILE ---
    try:
        # Use shutil.copy2() to copy the file (including metadata)
        # The file remains in source_path (origin_dir)
        shutil.copy2(source_path, destination_path)
        
        print(f"Successfully archived: {filename}")
        print(f"Original file saved in: {origin_dir}")
        print(f"Archived copy moved to: {destination_dir}")
        print(f"New archived filename: {new_filename}")

    except FileNotFoundError:
        print(f"Error: Source file not found at {source_path}")
    except Exception as e:
        print(f"An unexpected error occurred while archiving the file: {e}")

    print("")
    return






def main():


  with open("master\\schemas\\sections_order_info.json") as file:
    all_sections_info = json.load(file)
  with open("master\\schemas\\vcode_locs.json") as file:
    vcode_list = json.load(file)
  with open("master\\schemas\\misc_item_locs.json") as file:
    misc_list = json.load(file)

  for filename in [
    "sections_order_info.json",
    "misc_item_locs.json",
    "vcode_locs.json"]:
    move_and_archive_document(filename, "master\\schemas\\", "master\\schemas\\archive")

  master_list = pd.read_csv("master\\master_inventory_list.csv")

  deliverable_path = "deliverables\\printable_inventory_sheet.xlsx"


  deliverable = []

  all_sections_info

  for key in all_sections_info:
    section = all_sections_info[key]
    deliverable.append({
      "VENDOR/BRAND": key, 
      "ITEM_DESC": key,
      "UNIT": "",
      "PACK": "",
      "PER_PACK": "",
      "PRICE": "",
      "QUANTITY": ""
    })

    for item_info in section:
      master_key, misc_key = item_info
      # analyze vcode
      vendor, vcode = master_key.split(",", maxsplit=1)
      vendor = vendor.strip()
      vcode = vcode.strip()

      
      valid_master_key = vendor == "SYSCO" and vcode != "NAN"

      # sanitize vcode
      vcode = re.sub(r'\D', '', vcode)
      while len(vcode) < 7:
        # add leading zeros until such is not the case
        vcode = "".join(["0", vcode])

      # try to convert to int
      try:
        vcode = int(vcode)
      except Exception as e:
        # couldn't sanitize, forcing NAN
        vcode = "NAN"
        print(f"coudln't turn vcode {vcode} into an int: {e}")

      
      

      # check to see if master key already exists in master
      if valid_master_key and int(vcode) in master_list["VENDOR_CODE"].values:
        # get item information from master inventory list
        info = master_list[master_list['VENDOR_CODE'].values == int(vcode)]
        info["UNIT_TYPE"].values[0] = check_unit_type(info["UNIT_TYPE"].values[0])
        # there are some items in master_list without information
        # if this is the case for this item, instead
        # get and use info from vcode_locs to the best of our ability
        deliverable.append({
          "VENDOR/BRAND": "".join(
            [str(info["VENDOR"].values[0]), 
            "/", 
            str(info["BRAND"].values[0])]),
          "VENDOR_CODE": vcode, 
          "ITEM_DESC": info["ITEM"].values[0], 
          "UNIT": info["UNIT_TYPE"].values[0],
          "PACK": info["SUBUNIT"].values[0], 
          "PER_PACK": info["SUBUNIT_SIZE"].values[0], 
          "PRICE": info["UNIT_PRICE"].values[0], 
          "QUANTITY": ""
        })

      elif valid_master_key and master_key in vcode_list:
        # use item info from vcode_locs.json
        info = vcode_list[master_key]
        info["UNITS"][0] = check_unit_type(info["UNITS"][0])
        
        deliverable.append({
          "VENDOR/BRAND": vendor,
          "VENDOR_CODE": vcode,
          "ITEM_DESC": info["ITEM_DESC"][0], 
          "UNIT": info["UNITS"][0],
          "PACK": "", 
          "PER_PACK": "", 
          "PRICE": info["PRICES"][0],
          "QUANTITY": ""
        })
          
        # error, valid key but no information found with key
      elif valid_master_key:
        deliverable.append({
          "VENDOR/BRAND": vendor,
          "VENDOR_CODE": "",
          "ITEM_DESC": f"ITEM INFORMATION NOT FOUND {vcode}", 
          "UNIT": "",
          "PACK": "", 
          "PER_PACK": "", 
          "PRICE": "",
          "QUANTITY": ""
        })



      # master key not valid, using misc_key now
      else:
        # use the item information in misc_item_locs.json
        vendor, item_desc = misc_key.split(",", maxsplit = 1)
        item_desc = item_desc.strip()
        if vendor in ["NAN", "?"]:
          vendor = ""
        
        if misc_key in misc_list:
          
          info = misc_list[misc_key]
          info["UNITS"][0] = check_unit_type(info["UNITS"][0])

          deliverable.append({
            "VENDOR/BRAND": vendor,
            "VENDOR_CODE": "",
            "ITEM_DESC": item_desc, 
            "UNIT": info["UNITS"][0],
            "PACK": "", 
            "PER_PACK": "", 
            "PRICE": info["PRICES"][0] ,
            "QUANTITY": ""
          })

        else:
          deliverable.append({
            "VENDOR/BRAND": vendor,
            "VENDOR_CODE": "",
            "ITEM_DESC": item_desc, 
            "UNIT": "",
            "PACK": "", 
            "PER_PACK": "", 
            "PRICE": "",
            "QUANTITY": ""
          })
          











  deliverable = pd.DataFrame(
    deliverable, 
    columns = [
      "VENDOR/BRAND", "VENDOR_CODE", "ITEM_DESC", "UNIT", 
      "PACK", "PER_PACK", "PRICE", "QUANTITY"]
  )
  deliverable["EST_PRICE"] = 0
  deliverable["TOTAL EST VALUE"] = np.nan

  deliverable.to_excel(deliverable_path)

  # read in the file and do the excel formatting
  printed = pyxl.load_workbook(deliverable_path)

  ws = printed.active


  ## ADDING 10 BLANK ROWS AT BOTTOM OF SECTIONS
  n_blanks = 10

  total_rows = []
  for i, row in enumerate(ws.iter_rows(min_row = 2), start = 2):
    for cell in row:
      if "TOTAL:" in str(cell.value):
        total_rows.append(i)
        break

  for i in reversed(total_rows):
    ws.insert_rows(i, amount = n_blanks)


  ## CREATING ESTIMATED PRICE COLUMN AS PRODUCT OF QUANTITY AND PRICE
  header_row = 1
  headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}

  price_col = pyxl.utils.get_column_letter(headers["PRICE"])
  qty_col = pyxl.utils.get_column_letter(headers["QUANTITY"])
  est_col = pyxl.utils.get_column_letter(headers["EST_PRICE"])

  for row in range(header_row + 1, ws.max_row + 1):
    ws[f"{est_col}{row}"] = f"={price_col}{row}*{qty_col}{row}"


  ## TEXT WRAPPING FOR VENDOR/BRAND AND ITEM_DESC
  wrap_alignment = pyxl.styles.Alignment(wrap_text = True, vertical = "center")

  for col in ws.iter_cols():
    for cell in col:
      cell.alignment = wrap_alignment

  for cell in ws[1]:
    cell.alignment = wrap_alignment


  ## CENTER ALIGNMENT FOR VENDOR_CODE, UNIT, PACK, PER_PACK
  center_alignment = pyxl.styles.Alignment(
    horizontal = "center", vertical = "center")
  for col_name in ["VENDOR_CODE", "UNIT","PACK","PER_PACK"]:
    col_letter = pyxl.utils.get_column_letter(headers[col_name])
    for cell in ws[col_letter]:
      cell.alignment = center_alignment


  ## FREEZE TOP ROW AND PRINT HEADER ROW ON EACH PAGE
  ws.freeze_panes = "A2"
  ws.print_title_rows = "1:1"


  ## BOLD & ENLARGE FONT & HIGHLIGHT & MERGE SECTION HEADERS

  section_fill = pyxl.styles.PatternFill(
    start_color = "CCE5FF", end_color = "CCE5FF", fill_type = "solid"
  )
  section_font = pyxl.styles.Font(bold = True, size = 24)
  vender_col = headers['VENDOR/BRAND']

  for i, row in enumerate(ws.iter_rows(min_row = 2, max_col = ws.max_column), start = 2):
    cell = ws.cell(row = i, column = vender_col)
    if cell.value in all_sections_info.keys():
      # merge all 8 rows
      ws.merge_cells(f"B{i}:I{i}") 

      # format first leftmost visible cell in merged range
      merged_cell = ws.cell(row = i, column = 2)
      merged_cell.font = section_font
      merged_cell.fill = section_fill
      merged_cell.alignment = pyxl.styles.Alignment(
        horizontal = "center", vertical = "center"
      )



  ## HIGHLIGHT ROWS CONTAINING TOTAL:
  ## MERGE ROWS CONTAINING TOTAL:
  ## ADD NEW COLUMN TOTAL EST VALUE WITH SECTION SUMS
  total_fill = pyxl.styles.PatternFill(
    start_color = "FFFACD", end_color = "FFFACD", fill_type="solid"
  )
  last_total_row = 2

  for i, row in enumerate(ws.iter_rows(min_row = 2), start = 2):
    total_text = None
    for cell in row:
      if "TOTAL:" in str(cell.value):
        total_text = str(cell.value)
        break
    if total_text:
      # merging columns B - I for this row
      ws.merge_cells(f"B{i}:I{i}") 

      top_left = ws[f"B{i}"]
      top_left.value = total_text

      # determine the excel range for EST_PRICE (column I = 9)
      start_row = last_total_row + 1
      end_row = i - 1
      if end_row >= start_row:
        formula = f"=SUM(I{start_row}:I{end_row})"
      else:
        formula = "=0"

      total_cell = ws[f"J{i}"]
      total_cell.value = formula
      total_cell.number_format = pyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE



      # highlight & format the total row
      for cell in row:
        cell.fill = total_fill
        cell.font = pyxl.styles.Font(bold = True)
        cell.alignment = pyxl.styles.Alignment(
          horizontal = "center", vertical = "center")

      last_total_row = i


  ## ADD BORDER AROUND EACH ITEM
  thin_border = pyxl.styles.Border(
    left = pyxl.styles.Side(style = 'thin', color = '000000'),
    right = pyxl.styles.Side(style = 'thin', color = '000000'),
    top = pyxl.styles.Side(style = 'thin', color = '000000'),
    bottom = pyxl.styles.Side(style = 'thin', color = '000000')
  )

  for row in ws.iter_rows(
    min_row = 1, max_row = ws.max_row,
    min_col = 1, max_col = ws.max_column):
    for cell in row:
      cell.border = thin_border



  ## FORMATTING PRICE, EST_PRICE, TOTAL EST VALUE AS ACCOUNTING COLS
  for col in [headers["PRICE"], headers["EST_PRICE"]]:
    col_letter = pyxl.utils.get_column_letter(col)
    for cell in ws[col_letter]:
      cell.number_format = pyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

  ## FORMATTING QUANTITY TO BE NUMERIC COLUMN
  col_letter = pyxl.utils.get_column_letter(headers["QUANTITY"])
  for cell in ws[col_letter]:
    cell.number_format = pyxl.styles.numbers.FORMAT_NUMBER

  ## FORMATTING VENDOR_CODE TO BE TEXT TO PRESERVE LEADING ZEROS



  ## ADJUST COLUMN WIDTHS AND HEADER ROW HEIGHT
  ws.column_dimensions["B"].width = 20
  ws.column_dimensions["C"].width = 15
  ws.column_dimensions["D"].width = 15
  for col in range(5, 11):
    col_letter = pyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 10

  ws.row_dimensions[1].height = 35

  printed.save(deliverable_path)



  














  return


if __name__ == "__main__":
  main()
